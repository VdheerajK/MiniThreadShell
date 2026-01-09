import tkinter as tk
import psutil
import threading
import subprocess
import platform
from tkinter import messagebox
from subprocess import Popen, PIPE
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from collections import Counter
import time
import matplotlib.pyplot as plt
import boto3
from botocore.exceptions import NoCredentialsError



command_history = []


# Password (Change it here and use the same to login into the shell)
CORRECT_PASSWORD = "dheeraj284"  # You can change this to any password you like

# Logging the commands into Excel
def log_to_excel(command, output):
    filename = "command_log.xlsx"

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Command History"
        ws.append(["Timestamp", "Command", "Output"])
    else:
        wb = load_workbook(filename)
        ws = wb.active

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([timestamp, command, output.strip()])
    wb.save(filename)

# function to upload the excel log file to S3 bucket
def upload_to_s3(file_name, bucket_name, s3_file_name=None):
    if s3_file_name is None:
        s3_file_name = file_name

    try:
        s3 = boto3.client("s3", aws_access_key_id="*************", aws_secret_access_key="*******************")
        s3.upload_file(file_name, bucket_name, s3_file_name)
        print(f"{file_name} uploaded to S3 bucket: {bucket_name}")
    except FileNotFoundError:
        print(f" Error: {file_name} not found!")
    except NoCredentialsError:
        print("Error: AWS credentials not available.")
    except Exception as e:
        print(f"Upload failed: {e}")


def open_excel_file():
    filename = "command_log.xlsx"
    if not os.path.exists(filename):
        messagebox.showinfo("Info", "No log file found yet!")
        return

    try:
        system = platform.system()
        if system == "Darwin":  # macOS
            subprocess.call(["open", filename])
        elif system == "Windows":
            os.startfile(filename)
        elif system == "Linux":
            subprocess.call(["xdg-open", filename])
        else:
            messagebox.showerror("Error", "Unsupported OS")
    except Exception as e:
        messagebox.showerror("Error", f"Could not open file:\n{e}")

def start_system_monitor(frame):
    def update_stats():
        while True:
            cpu = psutil.cpu_percent()
            ram = psutil.virtual_memory().percent
            disk = psutil.disk_usage('/').percent

            cpu_label.config(text=f"CPU Usage: {cpu}%")
            ram_label.config(text=f"RAM Usage: {ram}%")
            disk_label.config(text=f"Disk Usage: {disk}%")

            time.sleep(1)

    import time
    global cpu_label, ram_label, disk_label

    cpu_label = tk.Label(frame, text="CPU Usage: --%", font=("Helvetica", 10), bg="#e0e0e0")
    ram_label = tk.Label(frame, text="RAM Usage: --%", font=("Helvetica", 10), bg="#e0e0e0")
    disk_label = tk.Label(frame, text="Disk Usage: --%", font=("Helvetica", 10), bg="#e0e0e0")

    cpu_label.pack(pady=2)
    ram_label.pack(pady=2)
    disk_label.pack(pady=2)

    # Run monitor in background
    threading.Thread(target=update_stats, daemon=True).start()

def show_shell_gui():
    shell_window = tk.Toplevel()
    shell_window.title("MiniThreadShell GUI")
    shell_window.configure(bg="#add8e6")
    shell_window.resizable(True, True)

    def run_command():
        cmd = entry.get()
        process = Popen(["./shell"], stdin=PIPE, stdout=PIPE, stderr=PIPE, text=True)
        output, error = process.communicate(input=f"{cmd}\nexit\n")
        result_text.delete("1.0", tk.END)
        result_text.insert(tk.END, f"$ {cmd}\n", "command")
        result_text.insert(tk.END, output, "output")
        result_text.insert(tk.END, error, "error")

  
        log_to_excel(cmd, output + error)

        command_history.append(cmd)

        upload_to_s3("command_log.xlsx", "dheerajprojectone")


    
    main_frame = tk.Frame(shell_window, bg="#add8e6")
    main_frame.pack(padx=10, pady=10)

    sidebar = tk.Frame(main_frame, width=150, bg="#e0e0e0", relief="groove", bd=2)
    sidebar.pack(side="left", fill="y", padx=(0, 10))

    start_system_monitor(sidebar)
    
    content_frame = tk.Frame(main_frame, bg="#add8e6")
    content_frame.pack(side="left", fill="both", expand=True)

    label = tk.Label(content_frame, text="Enter Command:", font=("Helvetica", 12), bg="#add8e6")
    label.pack(pady=(10, 0))

    entry = tk.Entry(content_frame, width=50, font=("Helvetica", 12))
    entry.pack(pady=5)

    run_btn = tk.Button(content_frame, text="Run", command=run_command, bg="#4CAF50", fg="black", font=("Helvetica", 11))
    run_btn.pack(pady=5)

    open_log_btn = tk.Button(content_frame, text="Open Excel Log", command=open_excel_file, bg="#FF9800", fg="black", font=("Helvetica", 11))
    open_log_btn.pack(pady=5)

    chart_btn = tk.Button(content_frame, text="Show Stats Chart", command=show_stats_chart, bg="#2196F3", fg="black", font=("Helvetica", 11))
    chart_btn.pack(pady=5)

    result_text = tk.Text(content_frame, height=15, width=60, bg="black", fg="black", font=("Courier", 10))
    result_text.pack(pady=10)

    result_text.tag_configure("command", foreground="cyan", font=("Courier", 10, "bold"))
    result_text.tag_configure("output", foreground="lime")
    result_text.tag_configure("error", foreground="red")


def check_password():
    entered = password_entry.get()
    if entered == CORRECT_PASSWORD:
        login_window.destroy()
        show_shell_gui()
    else:
        messagebox.showerror("Access Denied", "Incorrect password!")

from collections import Counter

def show_stats_chart():
    if not command_history:
        messagebox.showinfo("No Data", "No commands to display!")
        return

    top = Counter(command_history).most_common(5)
    commands = [cmd for cmd, count in top]
    counts = [count for cmd, count in top]

    plt.figure(figsize=(8, 5))
    bars = plt.bar(commands, counts, color='mediumseagreen')
    plt.title("Top 5 Most Used Commands")
    plt.xlabel("Command")
    plt.ylabel("Frequency")
    plt.ylim(0, max(counts) + 1)


    for bar in bars:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2, yval + 0.1, str(yval), ha='center')

    plt.tight_layout()
    plt.show()



root = tk.Tk()
root.withdraw()  

login_window = tk.Toplevel()
login_window.title("MiniThreadShell Login")
login_window.geometry("300x150")
login_window.configure(bg="#add8e6")
login_window.resizable(False, False)

login_label = tk.Label(login_window, text="Enter Password:", font=("Helvetica", 12), bg="black")
login_label.pack(pady=(20, 5))

password_entry = tk.Entry(login_window, show="*", font=("Helvetica", 12), width=25)
password_entry.pack(pady=5)
password_entry.focus()

login_btn = tk.Button(login_window, text="Login", command=check_password, bg="black", font=("Helvetica", 11))
login_btn.pack(pady=10)

root.mainloop()
